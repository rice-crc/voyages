# -*- coding: utf-8 -*-
from __future__ import (absolute_import, division, print_function,
                        unicode_literals)

import csv
import json
import logging
import time
import traceback
import urllib.error
import urllib.parse
import urllib.request
from builtins import map, range, str
from datetime import date
from itertools import (accumulate, filterfalse, groupby, islice, starmap,
                       takewhile)
from os import listdir, stat
from stat import ST_MTIME, ST_SIZE

import xlwt
from django.conf import settings
from django.contrib.admin.views.decorators import staff_member_required
from django.core.paginator import Paginator
from django.core.urlresolvers import reverse
from django.http import Http404, HttpResponse, HttpResponseRedirect
from django.shortcuts import redirect, render
from django.template import TemplateDoesNotExist, loader
from django.utils.translation import ugettext as _
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.gzip import gzip_page
from future import standard_library
from openpyxl import Workbook
from past.utils import old_div

import unidecode
from haystack.query import SearchQuerySet
from hurry.filesize import size
from voyages.apps.assessment.globals import get_map_year
from voyages.apps.common.export import download_xls
from voyages.apps.common.models import (SavedQuery,
                                        get_pks_from_haystack_results)
from voyages.apps.resources.models import Image
from voyages.apps.voyage.maps import VoyageRoutesCache

from .cache import CachedGeo, VoyageCache
from .forms import (GraphRemovePlotForm, GraphSelectionForm,
                    ResultsPerPageOptionForm, SimpleDateSearchForm,
                    SimpleNumericSearchForm, SimplePlaceSearchForm,
                    SimpleSelectBooleanForm, SimpleSelectSearchForm,
                    SimpleTextForm, TableSelectionForm,
                    TimeFrameSpanSearchForm, TimelineVariableForm,
                    UploadFileForm, graphs)
from .globals import (additional_var_dict, basic_variables,
                      calculate_maxmin_years, default_prettifier,
                      default_result_columns, display_methods,
                      display_methods_xls, display_unmangle_methods,
                      double_functions, general_variables, list_boolean_fields,
                      list_date_fields, list_imputed_nationality_values,
                      list_months, list_numeric_fields, list_place_fields,
                      list_select_fields, list_text_fields, no_mangle,
                      parameter_unmangle_methods, paginator_range_factors,
                      search_mangle_methods, session_expire_minutes,
                      summary_statistics, summary_statistics_columns,
                      table_columns, table_functions, table_rows,
                      var_dict, voyage_timeline_variables)
from .graphs import get_graph_data, graphs_y_axes
from .models import (Nationality, OwnerOutcome, ParticularOutcome, Resistance,
                     RigOfVessel, SlavesOutcome, VesselCapturedOutcome, Voyage)
from .search_indexes import ok_to_show_animation, VoyageIndex

standard_library.install_aliases()

# Here we enumerate all fields that should be cleared
# from the session if a reset is required.
reset_fields = [
    'voyages_tables_columns',
    'voyages_tables_rows',
    'voyages_tables_cells',
    'voyages_tables_omit',
    'voyage_timeline_form_option',
    'selected_graphs_tab',
    'tab_graphs_bar_defs',
    'tab_graphs_lin_defs',
    'tab_graphs_pie_defs',
    'tab_graphs_bar_defs_x_ind',
    'tab_graphs_bar_defs_y_ind',
    'tab_graphs_lin_defs_x_ind',
    'tab_graphs_lin_defs_y_ind',
    'tab_graphs_pie_defs_x_ind',
    'tab_graphs_pie_defs_y_ind',
]


def get_voyages_search_query_set():
    return SearchQuerySet().models(Voyage).filter(var_dataset=0)


def get_page(request, chapternum, sectionnum, pagenum):
    """
    Voyage Understanding the Database part

    Display an html page corresponding to the chapter-section-page

    Further content is rendered using the pagepath parameter
    """
    # We might want to do some error checking for pagenum here. Even though 404
    # will be raised if needed
    pagepath = "voyage/c" + chapternum + "_s" + \
        sectionnum + "_p" + pagenum + ".html"
    templatename = "voyage/c" + chapternum + \
        "_s" + sectionnum + "_generic" + ".html"

    try:
        loader.get_template(pagepath)
        loader.get_template(templatename)
        return render(request, templatename, dictionary={"pagepath": pagepath})
    except TemplateDoesNotExist:
        raise Http404


@staff_member_required
def download_file(request):
    """
    This view serves uploading files, which will be in
    the download section. It uses UploadFileForm to maintain
    information regarding uploaded files and call
    handle_uploaded_file() to store files on the disk.
    This view is available only for admin users.
    """
    templatename = 'upload.html'

    if request.method == 'POST':
        form = UploadFileForm(request.POST, request.FILES)
        if form.is_valid():
            handle_uploaded_file(request.FILES['downloadfile'])
            return HttpResponseRedirect('/admin/downloads')
    else:
        form = UploadFileForm()
    uploaded_files = listdir(settings.MEDIA_ROOT + '/download')
    uploaded_files_info = []
    for f in uploaded_files:
        st = stat(settings.MEDIA_ROOT + '/download/' + f)
        uploaded_files_info.append({
            'name': f,
            'size': size(st[ST_SIZE]),
            'date_mod': time.asctime(time.localtime(st[ST_MTIME]))
        })

    return render(request, templatename, {
        'form': form,
        'uploaded_files': uploaded_files_info
    })


def handle_uploaded_file(f):
    """
    Function handles uploaded files by saving them
    by chunks in the MEDIA_ROOT/download directory
    """
    with open('%s/%s/%s' % (settings.MEDIA_ROOT, 'download', f.name),
              'wb+') as destination:
        for chunk in f.chunks():
            destination.write(chunk)


def create_query_forms():
    """
    Uses the list of variables in globals.py and creates a form
    for each of them that is either a basic or a general variable
    Returns a list of dictionaries containing the var_name, var_full_name, form
    """
    voyage_span_first_year, voyage_span_last_year = calculate_maxmin_years(
    )
    form_list = []
    # for all basic and/or general variables
    for var in [
            x for x in var_dict if x['is_general'] or x['is_basic']
    ]:
        varname = var['var_name']
        vname = {'var_name': varname, 'var_full_name': var['var_full_name']}
        if varname in list_text_fields:
            form = SimpleTextForm(prefix=varname)
        elif varname in list_select_fields:
            choices = get_choices(varname)
            form = SimpleSelectSearchForm(prefix=varname)
            form.fields['choice_field'].choices = choices
        elif varname in list_numeric_fields:
            form = SimpleNumericSearchForm(initial={'options': '4'},
                                           prefix=varname)
        elif varname in list_date_fields:
            form = SimpleDateSearchForm(initial={
                'options': '1',
                'from_year': voyage_span_first_year,
                'to_year': voyage_span_last_year
            }, prefix=varname)
        elif varname in list_place_fields:
            nested_choices, flat_choices = get_nested_list_places(
                varname, var['choices'])
            form = SimplePlaceSearchForm(prefix=varname)
            form.fields['choice_field'].choices = flat_choices
            vname['nested_choices'] = nested_choices
            vname['selected_choices'] = []
        elif varname in list_boolean_fields:
            form = SimpleSelectBooleanForm(prefix=varname)
        else:
            pass
        form.fields['var_name_field'].initial = varname
        vname['form'] = form
        form_list.append(vname)
    return form_list


def retrieve_post_search_forms(post):
    """
    Retrieves the forms in the post and returns a list of dictionaries with
    var_name, var_full_name, and form
    """
    form_list = []
    # for all basic and/or general variables
    for var in [
            x for x in var_dict if x['is_general'] or x['is_basic']
    ]:
        varname = var['var_name']
        vname = {'var_name': varname, 'var_full_name': var['var_full_name']}
        form = None
        if varname in list_text_fields:
            form = SimpleTextForm(post, prefix=varname)
        elif varname in list_select_fields:
            form = SimpleSelectSearchForm(data=post, prefix=varname)
            form.fields['choice_field'].choices = get_choices(varname)
        elif varname in list_numeric_fields:
            form = SimpleNumericSearchForm(post, prefix=varname)
        elif varname in list_date_fields:
            form = SimpleDateSearchForm(post, prefix=varname)
        elif varname in list_boolean_fields:
            form = SimpleSelectBooleanForm(post, prefix=varname)
        elif varname in list_place_fields:
            form = SimplePlaceSearchForm(post, prefix=varname)
            nested_choices, flat_choices = get_nested_list_places(
                varname, var['choices'])
            form.fields['choice_field'].choices = flat_choices
            selected_choices = []
            if form.is_valid():
                selected_choices = [
                    int(i) for i in form.cleaned_data['choice_field']
                ]
            # Get the nested list places again to fill out the selected regions
            # and areas. Needed to get the flat choices and set the choice
            # field to that before form would validate.
            nested_choices, flat_choices = get_nested_list_places(
                varname, var['choices'], selected_choices)
            vname['nested_choices'] = nested_choices
            vname['selected_choices'] = selected_choices

        vname['form'] = form
        form_list.append(vname)
    return form_list


def create_forms_from_var_list(var_list):
    """
    Creates filled out forms based on a var_list
    """
    form_list = []
    vl = var_list.get('used_variable_names', '')
    vs = []
    if len(vl) > 0:
        vs = vl.split(';')
    for idx, varname in enumerate(vs):
        var = search_var_dict(varname)
        vname = {'var_name': varname, 'var_full_name': var['var_full_name']}
        if varname in list_text_fields:
            form = SimpleTextForm(prefix=varname)
            form.fields[
                'text_search'].initial = var_list[varname + '_text_search']
        elif varname in list_select_fields:
            choices = get_choices(varname)
            form = SimpleSelectSearchForm(prefix=varname)
            form.fields['choice_field'].choices = choices
            form.fields['choice_field'].initial = var_list[
                varname + '_choice_field'].split(';')
        elif varname in list_numeric_fields:
            form = SimpleNumericSearchForm(prefix=varname)
            opt = var_list[varname + '_options']
            form.fields['options'].initial = opt
            if opt == '1':  # Between
                form.fields[
                    'lower_bound'].initial = var_list[varname + '_lower_bound']
                form.fields[
                    'upper_bound'].initial = var_list[varname + '_upper_bound']
            elif opt == '2':  # Less than or equal to
                form.fields[
                    'threshold'].initial = var_list[varname + '_threshold']
            elif opt == '3':  # Greater than or equal to
                form.fields[
                    'threshold'].initial = var_list[varname + '_threshold']
            elif opt == '4':  # Equal to
                form.fields[
                    'threshold'].initial = var_list[varname + '_threshold']
        elif varname in list_date_fields:
            form = SimpleDateSearchForm(prefix=varname)
            form.fields['months'].initial = [
                str(x).zfill(2)
                for x in var_list[varname + '_months'].split(',')
            ]

            opt = var_list[varname + '_options']
            form.fields['options'].initial = opt
            if opt == '1':  # Between
                form.fields[
                    'from_year'].initial = var_list[varname + '_from_year']
                form.fields[
                    'from_month'].initial = var_list[varname + '_from_month']
                form.fields['to_year'].initial = var_list[varname + '_to_year']
                form.fields[
                    'to_month'].initial = var_list[varname + '_to_month']
            elif opt == '2':  # Less than or equal to
                form.fields['threshold_year'].initial = var_list[
                    varname + '_threshold_year']
                form.fields['threshold_month'].initial = var_list[
                    varname + '_threshold_month']
            elif opt == '3':  # Greater than or equal to
                form.fields['threshold_year'].initial = var_list[
                    varname + '_threshold_year']
                form.fields['threshold_month'].initial = var_list[
                    varname + '_threshold_month']
            elif opt == '4':  # Equal to
                form.fields['threshold_year'].initial = var_list[
                    varname + '_threshold_year']
                form.fields['threshold_month'].initial = var_list[
                    varname + '_threshold_month']
        elif varname in list_place_fields:
            selected_choices = [
                int(i) for i in var_list[varname + '_choice_field'].split(';')
            ]
            form = SimplePlaceSearchForm(prefix=varname)
            nested_choices, flat_choices = get_nested_list_places(
                varname, var['choices'], selected_choices)
            form.fields['choice_field'].choices = flat_choices
            vname['nested_choices'] = nested_choices
            vname['selected_choices'] = selected_choices
            form.fields['choice_field'].initial = selected_choices
        elif varname in list_boolean_fields:
            form = SimpleSelectBooleanForm(prefix=varname)
            form.fields['choice_field'].initial = var_list[
                varname + '_choice_field'].split(';')
        else:
            pass
        form.fields['var_name_field'].initial = varname
        form.fields['is_shown_field'].initial = str(idx)
        vname['form'] = form
        form_list.append(vname)
    return form_list


def create_var_dict(query_forms, time_frame_form):
    """
    query_forms: list of dictionaries with var_name and value (also probably
    var_full_name, but I don't think I need to count on that) returns a
    dictionary of var names and values for that var
    """
    # Creates a query dict based on all the restrictions the user has made
    var_list = {}
    used_variables = []
    # Year Time Frame Search
    if time_frame_form.is_valid():
        var_list['time_span_from_year'] = time_frame_form.cleaned_data[
            'frame_from_year']
        var_list['time_span_to_year'] = time_frame_form.cleaned_data[
            'frame_to_year']
    for qryform in [
            x for x in query_forms
            if x['form'].is_valid() and x['form'].is_form_shown()
    ]:
        # qform = next((l for l in query_forms if l['varname'] == search_var),
        # None)
        varname = qryform['var_name']
        used_variables.append(varname)
        form = qryform['form']
        if varname in list_text_fields:
            var_list[varname + '_text_search'] = unidecode.unidecode(
                form.cleaned_data['text_search'])
        elif varname in list_select_fields:
            # I don't think I need to unidecode this for the url
            var_list[varname + '_choice_field'] = ';'.join(
                form.cleaned_data['choice_field'])
        elif varname in list_numeric_fields:
            opt = form.cleaned_data['options']
            var_list[varname + '_options'] = opt
            if opt == '1':  # Between
                var_list[varname + '_lower_bound'
                         ] = form.cleaned_data['lower_bound']
                var_list[varname + '_upper_bound'
                         ] = form.cleaned_data['upper_bound']
            elif opt == '2':  # Less than or equal to
                var_list[varname + '_threshold'
                         ] = form.cleaned_data['threshold']
            elif opt == '3':  # Greater than or equal to
                var_list[varname + '_threshold'
                         ] = form.cleaned_data['threshold']
            elif opt == '4':  # Equal to
                var_list[varname + '_threshold'
                         ] = form.cleaned_data['threshold']
        elif varname in list_date_fields:
            var_list[varname + '_months'] = ','.join(
                form.cleaned_data['months'])
            opt = form.cleaned_data['options']
            var_list[varname + '_options'] = opt
            if opt == '1':  # Between
                var_list[
                    varname + '_from_year'] = form.cleaned_data['from_year']
                var_list[varname + '_from_month'
                         ] = form.cleaned_data['from_month']
                var_list[varname + '_to_year'] = form.cleaned_data['to_year']
                var_list[varname + '_to_month'] = form.cleaned_data['to_month']
            elif opt == '2':  # Less than or equal to
                var_list[
                    varname + '_threshold_year'] = form.cleaned_data[
                        'threshold_year']
                var_list[
                    varname + '_threshold_month'] = form.cleaned_data[
                        'threshold_month']
            elif opt == '3':  # Greater than or equal to
                var_list[
                    varname + '_threshold_year'] = form.cleaned_data[
                        'threshold_year']
                var_list[
                    varname + '_threshold_month'] = form.cleaned_data[
                        'threshold_month']
            elif opt == '4':  # Equal to
                var_list[
                    varname + '_threshold_year'] = form.cleaned_data[
                        'threshold_year']
                var_list[
                    varname + '_threshold_month'] = form.cleaned_data[
                        'threshold_month']
        elif varname in list_place_fields:
            var_list[varname + '_choice_field'] = unidecode.unidecode(';'.join(
                form.cleaned_data['choice_field']))
        elif varname in list_boolean_fields:
            var_list[varname + '_choice_field'] = ';'.join(
                form.cleaned_data['choice_field'])

    if len(used_variables) > 0:
        var_list['used_variable_names'] = ';'.join(used_variables)
    # for var in var_list:
    #    var_list[var] = unidecode.unidecode(unicode(var_list[var]))

    return var_list


def create_query_dict(var_list):
    """
    query_forms: list of dictionaries with var_name and form (also probably
    var_full_name, but I don't think I need to count on that) returns a tuple
    of query dict and a dictionary of variable names with the value(s) it has
    """
    # Creates a query dict based on all the restrictions the user has made
    query_dict = {}
    # Year Time Frame Search
    # if time_span_name in var_list:
    if 'time_span_from_year' in var_list and 'time_span_to_year' in var_list:
        query_dict['var_imp_arrival_at_port_of_dis__range'] = [
            var_list['time_span_from_year'], var_list['time_span_to_year']
        ]
    vl = var_list.get('used_variable_names', '')
    vs = []
    if len(vl) > 0:
        vs = vl.split(';')
    for varname in vs:
        try:
            mangle_method = search_mangle_methods.get(
                varname, no_mangle)
            if varname == 'var_sources':
                query_dict[
                    "var_sources_plaintext_search__contains"] = mangle_method(
                        var_list[varname + '_text_search'])
            elif varname in list_text_fields:
                query_dict[varname + "__contains"] = mangle_method(
                    var_list[varname + '_text_search'])
            elif varname in list_select_fields:
                query_dict[varname + "_idnum" + "__in"] = [
                    int(i)
                    for i in mangle_method(
                        var_list[varname + '_choice_field']).split(';')
                    if i != ''
                ]
            elif varname in list_numeric_fields:
                opt = var_list[varname + '_options']
                if opt == '1':  # Between
                    query_dict[varname + "__range"] = [
                        mangle_method(var_list[varname + '_lower_bound']),
                        mangle_method(var_list[varname + '_upper_bound'])
                    ]
                elif opt == '2':  # Less than or equal to
                    query_dict[varname + "__lte"] = mangle_method(
                        var_list[varname + '_threshold'])
                elif opt == '3':  # Greater than or equal to
                    query_dict[varname + "__gte"] = mangle_method(
                        var_list[varname + '_threshold'])
                elif opt == '4':  # Equal to
                    query_dict[varname + "__exact"] = mangle_method(
                        var_list[varname + '_threshold'])
            elif varname in list_date_fields:
                month_list = varname + '_months'
                if month_list in var_list:
                    months = [
                        int(x) for x in var_list[month_list].split(',')
                    ]
                    # Only filter by months if not all the months are included
                    if len(months) < 12:
                        query_dict[
                            varname + '_month__in'] = [int(x) for x in months]
                opt = var_list[varname + '_options']
                if opt == '1':  # Between
                    to_date = None
                    var_month = var_list[f'{varname}_to_month']
                    var_year = var_list[f'{varname}_to_year']
                    if int(var_month) == 12:
                        to_date = format_date(
                            int(mangle_method(var_year)) + 1, 1)
                    else:
                        to_date = format_date(
                            int(mangle_method(var_year)),
                            int(mangle_method(var_month)) + 1)
                    query_dict[varname + "__range"] = [
                        format_date(
                            mangle_method(var_list[varname + '_from_year']),
                            mangle_method(var_list[varname + '_from_month'])),
                        to_date
                    ]
                elif opt == '2':  # Less than or equal to
                    month = var_list[varname + '_threshold_month']
                    year = var_list[varname + '_threshold_year']
                    to_date = None
                    if int(month) == 12:
                        to_date = format_date(int(mangle_method(year)) + 1, 1)
                    else:
                        to_date = format_date(
                            int(mangle_method(year)),
                            int(mangle_method(month)) + 1)
                    query_dict[varname + "__lte"] = to_date
                elif opt == '3':  # Greater than or equal to
                    month = var_list[varname + '_threshold_month']
                    year = var_list[varname + '_threshold_year']
                    to_date = format_date(mangle_method(year),
                                          mangle_method(month))
                    query_dict[varname + "_gte"] = to_date
                elif opt == '4':  # In
                    to_date = None
                    month = var_list[varname + '_threshold_month']
                    year = var_list[varname + '_threshold_year']
                    if int(month) == 12:
                        to_date = format_date(int(mangle_method(year)) + 1, 1)
                    else:
                        to_date = format_date(
                            int(mangle_method(year)),
                            int(mangle_method(month)) + 1)
                    query_dict[varname + "__range"] = [
                        format_date(mangle_method(year), mangle_method(month)),
                        to_date]
            elif varname in list_place_fields:
                query_dict[varname + "_idnum" + "__in"] = [
                    int(i)
                    for i in mangle_method(var_list[
                        varname + '_choice_field']).split(';')
                    if i != ''
                ]
            elif varname in list_boolean_fields:
                query_dict[varname + "__in"] = mangle_method(
                    var_list[varname + '_choice_field']).split(';')
        except Exception:
            print(f"Failure when mangling variable {varname}. "
                  "It will be removed from the search.")
            traceback.print_exc()
    return query_dict


def create_var_list_from_url(get):
    """
    Takes a request.GET and returns a var_list with the search parameters
    """
    var_list = {}
    for i in get:
        var_list[i] = get.get(i)
    return var_list


# Takes a var list and then gives tuples of strings that describe the query in
# a nice way Used for formatting the display of previous_queries


def prettify_var_list(varlist):
    output = []
    qdict = create_query_dict(varlist)
    # For some reason, when time_span is set, it also shows "Year arrived with
    # slaves*"
    if 'time_span_from_year' in varlist and 'time_span_to_year' in varlist:
        beg = varlist['time_span_from_year']
        end = varlist['time_span_to_year']
        output.append((_('Time frame:'), str(beg) + ' - ' + str(end)))
    for kvar, vvar in list(qdict.items()):
        varname = kvar.split('__')[0]
        is_real_var = False
        fullname = ''
        accepted_suffixes = varname.endswith(('_idnum', '_plaintext_search'))
        for var in var_dict:
            if varname == var['var_name'] or (
                    accepted_suffixes and varname.startswith(var['var_name'])):
                fullname = var['var_full_name']
                is_real_var = True
                break
        if not is_real_var:
            # it is a month variable
            varn = varname[:-6]
            for var in var_dict:
                if varn == var['var_name']:
                    fullname = var['var_full_name']
                    break
            month_dict = {}
            for monnum, monval in list_months:
                month_dict[int(monnum)] = monval
            output.append((fullname + _(" month:"),
                           ', '.join([month_dict[int(i)] for i in vvar])))
            continue
        unmangle_method = parameter_unmangle_methods.get(
            varname, default_prettifier(varname))
        tvar = unmangle_method(vvar)
        if isinstance(tvar, (list, tuple)):
            value = str(u', '.join(map(str, tvar)))
        else:
            value = tvar
        prefix = ''
        if varname + '_options' in varlist:
            opt = varlist[varname + '_options']
            if opt == '1' and len(vvar) >= 2:
                viaapod = 'var_imp_arrival_at_port_of_dis'
                if varname == viaapod or varname not in list_date_fields:
                    beg = str(tvar[0])
                    end = str(tvar[1])
                else:
                    tod = None
                    if vvar[1].month == 1:
                        tod = date(vvar[1].year - 1, 12, vvar[1].day)
                    else:
                        tod = date(
                            vvar[1].year, vvar[1].month - 1, vvar[1].day)
                    beg = str(unmangle_method(vvar[0]))
                    end = str(unmangle_method(tod))
                value = _('between ') + beg + _(' and ') + end
            elif opt == '4':
                if isinstance(vvar, (list, tuple)):
                    value = _('in ') + str(unmangle_method(vvar[0]))
                else:
                    value = _('equal to ') + str(tvar)
            elif isinstance(vvar, (list, tuple)):
                continue
            elif opt == '2':
                if varname == 'var_imp_arrival_at_port_of_dis':
                    value = _('before ') + str(tvar)
                elif varname in list_date_fields:
                    tod = None
                    if vvar.month == 1:
                        tod = date(vvar.year - 1, 12, vvar.day)
                    else:
                        tod = date(vvar.year, vvar.month - 1, vvar.day)
                    value = _('before ') + str(unmangle_method(tod))
                else:
                    value = _('at most ') + str(tvar)
            elif opt == '3':
                if varname == 'var_imp_arrival_at_port_of_dis':
                    value = _('after ') + str(tvar)
                elif varname in list_date_fields:
                    value = _('after ') + str(tvar)
                else:
                    value = _('at least ') + str(tvar)
        # Prevent display of 'Year arrived with slaves*' when it is just the
        # time frame
        if not isinstance(vvar, (list, tuple)) or \
                varname not in list_numeric_fields or \
                (varname + '_options') in varlist:
            output.append((fullname + ":", (prefix + value)))
    return output


def first_match(items):
    return items[0] if len(items) > 0 else None


def voyage_map(request, voyage_id):
    """
    Displays the map for a voyage
    """
    voyage = first_match(
        get_voyages_search_query_set().filter(var_voyage_id=int(voyage_id)))
    if voyage:
        year_completed = int(
            voyage.var_imp_voyage_began) if voyage.var_imp_voyage_began else 0
        map_year = get_map_year(year_completed, year_completed)
    else:
        map_year = None
    return render(
        request, "voyage_info.html", {
            'tab': 'map',
            'map_year': map_year,
            'voyage_id': voyage_id,
            'voyage': voyage
        })


def voyage_images(request, voyage_id):
    """
    Displays the images for a voyage
    """
    voyage = first_match(
        get_voyages_search_query_set().filter(var_voyage_id=int(voyage_id)))
    images = []
    if voyage:
        images = list(Image.objects.filter(voyage=int(voyage_id)))
    return render(request, "voyage_info.html", {
        'tab': 'images',
        'voyage_id': voyage_id,
        'voyage': voyage,
        'images': images
    })


def voyage_variables_data(voyage_id, show_imputed=True):
    voyagenum = int(voyage_id)
    voyage = first_match(
        get_voyages_search_query_set().filter(var_voyage_id=voyagenum))
    if voyage is None:
        return None, []
    # Apply the matching method (if there is one) in the display_method_details
    # dict for each variable value in the voyage and return a dict of varname:
    # varvalue
    voyagevariables = voyage.get_stored_fields()
    # for vname, vvalue in voyage.get_stored_fields().items():
    #    voyagevariables[vname] = display_methods_details.get(vname,
    #    no_mangle)(vvalue, voyagenum)
    allvargroups = groupby(var_dict, key=lambda x: x['var_category'])
    allvars = []
    for i in allvargroups:
        group = i[0]
        glist = list(
            x for x in i[1]
            if show_imputed or not x['var_full_name'].endswith('*')
        )
        for idx, j in enumerate(glist):
            val = str("")
            if voyagevariables[j['var_name']]:
                mangle_method = display_unmangle_methods.get(
                    j['var_name'], default_prettifier(j['var_name']))
                val = str(
                    mangle_method(voyagevariables[j['var_name']], voyagenum))
            if val == u'[]':
                val = u''
            if idx == 0:
                # For the first variable, give the number of variables in the
                # group, and give the name of the group as a tuple in the first
                # entry of the triple for the row
                newvar = (len(glist), str(group))
            else:
                newvar = (None, None)
            allvars.append((newvar,
                            str(j['var_full_name']), val, j['var_name']))
    return voyage, allvars


def voyage_variables(request, voyage_id):
    """
    Displays all the variables for a single voyage
    """
    (voyage, allvars) = voyage_variables_data(voyage_id)

    return render(
        request, "voyage_info.html", {
            'voyage_variables': allvars,
            'voyage': voyage,
            'tab': 'variables',
            'voyage_id': voyage_id
        })


def reload_cache(_):
    VoyageCache.load(True)
    return HttpResponse("Voyages cache reloaded")


@ csrf_exempt
@ gzip_page
def search(request):
    """
    Handles the Search the Database part
    """
    no_result = False
    query_dict = {}
    result_data = {}
    tab = 'result'
    result_data['summary_statistics_columns'] = summary_statistics_columns
    form_list = []
    voyage_span_first_year, voyage_span_last_year = calculate_maxmin_years()
    search_url = None
    results = None
    time_frame_form = None
    results_per_page_form = None
    results_per_page = 10
    basic_list_contracted = False
    previous_queries = {}
    collabels = []
    prev_queries_open = False
    row_list = []
    table_stats_form = None

    col_totals = []
    extra_cols = 0
    num_col_labels_before = 1
    num_col_labels_total = 1
    num_row_labels = 1
    graphs_select_form = None
    graph_data = None
    graph_xfun_index = None
    graphs_tab = None
    graph_remove_plots_form = None
    inline_graph_png = None

    # Timeline
    timeline_data = []
    timeline_form = None
    timeline_chart_settings = {}

    order_by_field = request.session.get('voyages_order_by_field',
                                         'var_voyage_id')
    sort_direction = request.session.get('voyages_sort_direction', 'asc')

    # Map
    map_year = '1750'

    # Check if we are restoring POST data from session,
    # which is what would happen when accessing a permalink.
    if 'voyages_post_data' in request.session:
        request.method = 'POST'
        request.POST = request.session.pop('voyages_post_data')

    # If there is no requested page number, serve 1
    current_page = 1
    desired_page = request.POST.get('desired_page')
    if desired_page:
        current_page = desired_page

    submit_val = request.POST.get('submitVal')

    if 'submit_val' in request.GET:
        submit_val = request.GET['submit_val']

    # If session has expired (no search activity for the last
    # session_expire_minutes time) then clear the previous queries
    old_time = request.session.get('last_access_time', 0.0)
    if old_time < time.time() - session_expire_minutes * 60.0:
        # request.session.clear()
        request.session['previous_queries'] = []

    request.session['last_access_time'] = time.time()

    # if used_variable_names or the pair of time_span_from_year and
    # time_span_to_year keys are in request.GET, then that means that it is a
    # query url and we should get the query from it. or if it is
    # restore_prev_query, then restore it from the session.
    if submit_val == 'restore_prev_query' or (
            request.method == "GET" and (
                'used_variable_names' in request.GET or (
                    'time_span_from_year' in request.GET and (
                        'time_span_to_year' in request.GET)))):
        # Search parameters were specified in the url
        var_list = {}
        results_per_page_form = ResultsPerPageOptionForm()
        if submit_val == 'restore_prev_query':
            qnum = int(
                request.POST.get('prev_query_num',
                                 request.GET.get('prev_query_num')))
            if 'prev_query_num' in request.GET:
                current_page = request.session.get('current_page', 0)
                results_per_page_form.fields[
                    'option'].initial = request.session.get(
                        'results_per_page_choice', '1')
                results_per_page = dict(
                    results_per_page_form.fields['option'].choices)[
                        request.session.get('results_per_page_choice', '1')]
            qprev = request.session.get('previous_queries', [])
            if 0 <= qnum < len(qprev):
                var_list = qprev[qnum]
                qprev.remove(qprev[qnum])
            request.session['previous_queries'] = qprev
            prev_queries_open = True
        else:
            var_list = create_var_list_from_url(request.GET)
        if 'previous_queries' not in request.session:
            request.session['previous_queries'] = []
        request.session['previous_queries'] = [
            var_list
        ] + request.session['previous_queries']
        form_list = create_query_forms()
        var_name_indexes = {}
        for idx, form in enumerate(form_list):
            var_name_indexes[form['var_name']] = idx
        filled_form_list = create_forms_from_var_list(var_list)
        to_remove_numbers = []
        for form in filled_form_list:
            to_remove_numbers.append(var_name_indexes[form['var_name']])
            form_list.append(form)
        for idx in sorted(to_remove_numbers, reverse=True):
            del form_list[idx]
        time_frame_form = TimeFrameSpanSearchForm(
            initial={
                'frame_from_year':
                    var_list.get('time_span_from_year',
                                 voyage_span_first_year),
                'frame_to_year':
                    var_list.get('time_span_to_year', voyage_span_last_year)
            })
        query_dict = create_query_dict(var_list)
        results = perform_search(query_dict, None, order_by_field,
                                 sort_direction, request.LANGUAGE_CODE)
        search_url = request.build_absolute_uri(reverse(
            'voyage:search',)) + "?" + urllib.parse.urlencode(var_list)

    elif request.method == "GET" or request.POST.get('submitVal') == 'reset':
        # A new search is being performed
        # Clear session keys.
        for key in reset_fields:
            request.session.pop(key, None)
        results_per_page_form = ResultsPerPageOptionForm()
        form_list = create_query_forms()
        time_frame_form = TimeFrameSpanSearchForm(
            initial={
                'frame_from_year': voyage_span_first_year,
                'frame_to_year': voyage_span_last_year
            })
        results = get_voyages_search_query_set().order_by('var_voyage_id')
        if request.POST.get('submitVal') == 'reset':
            request.session['result_columns'] = get_new_visible_attrs(
                default_result_columns)
    elif request.method == "POST":

        # A normal search is being performed, or it is on another tab, or it is
        # downloading a file
        rpp = 'results_per_page'
        rppc = 'results_per_page_choice'
        results_per_page_form = ResultsPerPageOptionForm(request.POST)
        if results_per_page_form.is_valid():
            results_per_page = results_per_page_form.cleaned_option()
            request.session[rppc] = results_per_page_form.cleaned_data[
                'option']
            request.session[rpp] = results_per_page
        elif rpp in request.session and rppc in request.session:
            results_per_page = request.session[rpp]
            results_per_page_form.fields['option'].initial = request.session[
                rppc]
            results_per_page_form = ResultsPerPageOptionForm(
                {u'option': request.session[rppc]})
        display_columns = request.session.get(
            'result_columns',
            get_new_visible_attrs(default_result_columns))

        ble = request.POST.get('basic_list_expanded')
        basic_list_contracted = not ble

        form_list = retrieve_post_search_forms(request.POST)
        time_frame_form = TimeFrameSpanSearchForm(request.POST)
        var_list = create_var_dict(form_list, time_frame_form)
        if 'previous_queries' not in request.session:
            request.session['previous_queries'] = []
        if submit_val != 'delete_prev_query':
            qprev = request.session['previous_queries']
            if len(qprev) < 1 or not qprev[0] == var_list:
                request.session['previous_queries'] = [var_list] + qprev
        search_url = request.build_absolute_uri(reverse(
            'voyage:search',)) + "?" + urllib.parse.urlencode(var_list)
        query_dict = create_query_dict(var_list)

        order_by_field = request.POST.get('order_by_field', order_by_field)
        sort_direction = request.POST.get('sort_direction', sort_direction)
        request.session['voyages_order_by_field'] = order_by_field
        request.session['voyages_sort_direction'] = sort_direction
        results = perform_search(query_dict, None, order_by_field,
                                 sort_direction, request.LANGUAGE_CODE)

        if submit_val == 'configColumn':
            tab = 'config_column'
        elif submit_val == 'applyConfig':
            request.session['result_columns'] = get_new_visible_attrs(
                request.POST.getlist('configure_visibleAttributes'))
            tab = 'result'
        elif submit_val == 'cancelConfig':
            tab = 'result'
        elif submit_val == 'restoreConfig':
            request.session['result_columns'] = get_new_visible_attrs(
                default_result_columns)
            tab = 'config_column'
        elif submit_val == 'tab_results':
            tab = 'result'
        elif submit_val == 'tab_statistics':
            tab = 'statistics'
            result_data['summary_statistics'] = retrieve_summary_stats(results)
        elif (submit_val and submit_val.startswith("tab_tables")
              ) or submit_val == 'xls_download_table':
            # row_cell_values is what is displayed in the cells in the table,
            # it is a list of triples which contain the row_label, the cell
            # values, then the row total
            # rowlabels is a list of lists of row label tuples (e.g. there is
            # the region and the port). Typically these will just be a list of
            # lists with one entry that is the label tuple for that row/
            # column labels is similar, but it is a list of column label lists,
            # and will typically be a list of one element that is a list of the
            # column label tuples entries in the rowlabels/collabels matrix are
            # tuples that contain the label and then the row/column span of
            # that cell.
            # Most of the time the row/column span will just be 1.
            xls_table = []
            tab = 'tables'
            pst = dict(request.POST.items())

            # Try to retrieve sessions values
            tables_columns = request.session.get('voyages_tables_columns')
            tables_rows = request.session.get('voyages_tables_rows')
            tables_cells = request.session.get('voyages_tables_cells')
            omit_empty = request.session.get('voyages_tables_omit')

            # Collect settings (if possible retrieve from tyhe session)
            if 'columns' not in pst:
                pst['columns'] = tables_columns or '7'
            if 'rows' not in pst:
                pst['rows'] = tables_rows or '12'
            if 'cells' not in pst:
                pst['cells'] = tables_cells or '1'
            omit_empty = ('omit_empty' in pst
                          if submit_val == "tab_tables_in"
                          else omit_empty or True)
            pst['omit_empty'] = omit_empty
            request.session['voyages_tables_omit'] = omit_empty

            # Update sessions with updated values
            request.session['voyages_tables_columns'] = pst['columns']
            request.session['voyages_tables_rows'] = pst['rows']
            request.session['voyages_tables_cells'] = pst['cells']

            table_stats_form = TableSelectionForm(pst)
            table_row_query_def = table_rows[12]
            table_col_query_def = table_columns[7]
            display_function = table_functions[1][1]
            display_fun_name = table_functions[1][0]
            if table_stats_form.is_valid():
                # If form is valid, collect all necessary settings: column, row
                # and cell variables
                table_row_query_def = table_rows[int(
                    table_stats_form.cleaned_data['rows'])]
                table_col_query_def = table_columns[int(
                    table_stats_form.cleaned_data['columns'])]
                display_function = table_functions[int(
                    table_stats_form.cleaned_data['cells'])][1]
                display_fun_name = table_functions[int(
                    table_stats_form.cleaned_data['cells'])][0]
                table_stats_form.omit_empty = omit_empty

            restrict_query = {}
            # Get the variable name of the variable used to filter the rows so
            # we can constrain the column totals to voyages with the row
            # variable defined
            table_row_var_name = ''
            if len(table_row_query_def[1]) > 0:
                # The query def is a triple with the 2nd element being a list
                # that list is a list of tuples with the label and the query
                # dict the query dict is a dictionary with 1 element which the
                # key is the var name with a '__' and then the query type (e.g.
                # "__exact")
                table_row_var_name = list(
                    table_row_query_def[1][0][1].keys())[0].split('__')[0]
            table_row_var = search_var_dict(table_row_var_name)
            if not table_row_var:
                for var in additional_var_dict:
                    if var['var_name'] == table_row_var_name:
                        table_row_var = var
                        break
            if table_row_var:
                if table_row_var['var_type'] == 'numeric':
                    restrict_query[table_row_var_name + "__gte"] = -1
                elif table_row_var['var_type'] == 'date':
                    restrict_query[
                        f"{table_row_var_name}__gte"] = date(1, 1, 1)
                else:
                    restrict_query[table_row_var_name + "__gte"] = ""
            elif table_row_var_name.endswith('_idnum'):
                restrict_query[table_row_var_name + "__gte"] = "-1"
            elif table_row_var_name != '':
                restrict_query[table_row_var_name + "__gte"] = ""

            # Get the variable name for the column
            table_col_var_name = ''
            if len(table_col_query_def[1]) > 0:
                table_col_var_name = list(
                    table_col_query_def[1][0].keys())[0].split('__')[0]
            table_col_var = search_var_dict(table_col_var_name)
            if not table_col_var:
                for var in additional_var_dict:
                    if var['var_name'] == table_col_var_name:
                        table_col_var = var
            qname = f"{table_col_var_name}_gte"
            if table_col_var:
                vtype = table_col_var['var_type']
                restrict_query[qname] = (
                    -1 if vtype == 'numeric' else
                    date(1, 1, 1) if vtype == 'date' else
                    "")
            elif table_col_var_name.endswith('_idnum'):
                restrict_query[qname] = "-1"
            elif table_col_var_name != '':
                restrict_query[qname] = ""

            tableresults = results.filter(**restrict_query)

            extra_cols = table_row_query_def[2]
            cell_values = []
            used_col_query_sets = []

            # Transform tuples to lists
            listed_cols = [list(map(list, k)) for k in table_col_query_def[2]]

            # Create column labels
            collabels = [list(i) for i in listed_cols]

            num_col_labels_total = len(collabels)
            num_row_labels = extra_cols + 1
            remove_cols = []
            is_double_fun = display_fun_name in double_functions

            for idx, colquery in enumerate(table_col_query_def[1]):
                colqueryset = tableresults.filter(**colquery)
                if omit_empty and colqueryset.count() == 0:
                    # Find column label that matches, then find the parent
                    # labels that match Generate the list of subcolumns for the
                    # parent column label
                    remove_cols.insert(0, idx)
                    continue
                if is_double_fun:
                    display_col_total = display_function(
                        colqueryset, None, colqueryset, tableresults)
                    col_totals.append(display_col_total[0])
                    col_totals.append(display_col_total[1])
                else:
                    col_totals.append(
                        display_function(colqueryset, None, colqueryset,
                                         tableresults))
                used_col_query_sets.append((colquery, colqueryset))
            for col in remove_cols:
                for idt, collbllist in enumerate(collabels):
                    enum_list = enumerate(collbllist)
                    idc, colstuff = next(islice(enum_list, len(list(
                        takewhile(lambda idy, col=col: idy <= col, accumulate(
                            starmap(lambda i, c: c[1], enum_list))))), None))
                    collabels[idt][idc] = [colstuff[0], colstuff[1] - 1]
            remove_rows = []
            if is_double_fun:
                collabels = [[[j, k * 2] for j, k in i] for i in collabels]
                lastcol = []
                for i in collabels[-1]:
                    lastcol.append(['Embarked', old_div(i[1], 2)])
                    lastcol.append(['Disembarked', old_div(i[1], 2)])
                collabels.append(lastcol)
            num_col_labels_before = len(collabels)
            xls_row = []
            for idx, i in enumerate(collabels):
                xls_row = []
                for j in range(num_row_labels):
                    xls_row.append('')
                for lbl, num in i:
                    if num == 0:
                        continue
                    xls_row.append(lbl)
                    xls_row.extend('' * (num - 1))
                xls_table.append(xls_row)
                if idx == 0:
                    if is_double_fun:
                        xls_row.append('Total Embarked')
                        xls_row.append('Total Disembarked')
                    else:
                        xls_row.append('Totals')

            for idx, rowstuff in enumerate(table_row_query_def[1]):
                xls_row = []
                rowlabels = rowstuff[0]
                rowquery = rowstuff[1]
                rowqueryset = tableresults.filter(**rowquery)
                if omit_empty and rowqueryset.count() == 0:
                    remove_rows.insert(0, idx)
                row_cell_values = []
                # Iterate through column labels to make the labels for the xls
                # download
                for colquery, colqueryset in used_col_query_sets:
                    cell_queryset = rowqueryset
                    if rowqueryset.count() > 0:
                        cell_queryset = rowqueryset.filter(**colquery)

                    display_result = display_function(
                        cell_queryset, rowqueryset, colqueryset,
                        tableresults)
                    if is_double_fun:
                        row_cell_values.append(display_result[0])
                        row_cell_values.append(display_result[1])
                        xls_row.append(display_result[0] or '')
                        xls_row.append(display_result[1] or '')
                    else:
                        row_cell_values.append(display_result)
                        xls_row.append(display_result or '')
                cell_values.append(row_cell_values)
                row_total = display_function(rowqueryset, rowqueryset, None,
                                             tableresults)
                row_list.append([
                    [(i[0], i[1]) for i in rowlabels],
                    row_cell_values,
                    row_total,
                ])
                if is_double_fun:
                    xls_row.append(row_total[0] or '')
                    xls_row.append(row_total[1] or '')
                else:
                    xls_row.append(row_total or '')
                xls_table.append(xls_row)
                # cell_displays.append((rowlbl, row_cell_displays, row_total))
            for rownum in remove_rows:
                xls_table.pop(rownum + num_col_labels_before)
                row_list[rownum] = ([
                    (i[0], i[1] - 1) for i in row_list[rownum][0]
                ], row_list[rownum][1], row_list[rownum][2])
                # Now find the rows with the headers for it and reduce those
                # header counts by 1
                for idx, rl in enumerate(row_list):
                    if idx >= rownum:
                        continue
                    limit = rownum - idx
                    rowlbl = list(rl[0])
                    decrements = list(filterfalse(
                        lambda idy, i, lim=limit: i[1] < lim,
                        enumerate(rowlbl)))
                    for idy, i in decrements:
                        # Don't handle the case for the rownum, since that has
                        # already been decremented
                        row_list[idx][0][idy][1] -= 1
                # Now handle the case when this row we are removing has headers
                # of its own
                rowlbl = list(row_list[rownum][0])
                rowlbl.reverse()
                for lbl, num in rowlbl:
                    if num > 0:
                        row_list[rownum + 1][0].insert(0, (lbl, num))
                row_list.pop(rownum)
            for idx, row in enumerate(row_list):
                rowlbl = list(row[0])
                for i in range(num_row_labels - len(rowlbl)):
                    xls_table[idx + num_col_labels_before].insert(0, '')
                rowlbl.reverse()
                for i in rowlbl:
                    xls_table[idx + num_col_labels_before].insert(
                        num_row_labels - len(rowlbl), i[0])
            if is_double_fun:
                grand_total_value = display_function(tableresults, None, None,
                                                     tableresults)
                col_totals.append(grand_total_value[0])
                col_totals.append(grand_total_value[1])
            else:
                col_totals.append(
                    display_function(tableresults, None, None, tableresults))
            xls_row = ['Totals'] + (num_row_labels - 1) * ['']
            for i in col_totals:
                xls_row.append(i or '')
            xls_table.append(xls_row)
            if submit_val == 'xls_download_table':
                response = HttpResponse(
                    content_type='application/'
                    'vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                )
                response[
                    'Content-Disposition'] = 'attachment; filename="data.xlsx"'
                wb = Workbook()
                ws = wb.active
                for idx, i in enumerate(xls_table):
                    for idy, j in enumerate(i):
                        ws.cell(row=idx + 1, column=idy + 1).value = j
                wb.save(response)
                return response

            # If empty, check if any columns need to be removed
            # TODO: Temporarily commented (waiting for David's response.
            # if omit_empty:
            #    remove_empty_columns(0, row_list, collabels, col_totals)

        elif submit_val and submit_val.startswith('tab_graphs'):
            tab = 'graphs'
            # Each tab name is encoded by 3 characters (lin, bar, pie)
            # representing the types of graphs supported.
            graphs_tab = submit_val[:len('tab_graphs_???')]
            if graphs_tab not in [
                    'tab_graphs_lin', 'tab_graphs_bar', 'tab_graphs_pie'
            ]:
                graphs_tab = request.session.get('selected_graphs_tab',
                                                 'tab_graphs_lin')
            request.session['selected_graphs_tab'] = graphs_tab
            session_defs_key = graphs_tab + '_defs'
            # Default value for x-axis or session stored value (if any).
            xind = request.session.get(session_defs_key + '_x_ind', 0)
            yind = request.session.get(session_defs_key + '_y_ind', 0)
            # Handle adding a y-function or simply changing the x-function.
            xfuns = (graphs.graphs_x_axes
                     if graphs_tab == 'tab_graphs_lin'
                     else graphs.other_graphs_x_axes)
            graphs_form_params = {
                'data': request.POST,
                'prefix': graphs_tab,
                'xfunctions': xfuns
            }
            if graphs_tab == 'tab_graphs_pie':
                graphs_form_params['xfield_label'] = 'Sectors'
                graphs_form_params['yfield_label'] = 'Values'
            graphs_select_form = GraphSelectionForm(**graphs_form_params)
            if graphs_select_form.is_valid():
                # Form is valid and we should use the xy-axes specified by it.
                xind = int(graphs_select_form.cleaned_data['xselect'])
                yind = int(graphs_select_form.cleaned_data['yselect'])
            else:
                graphs_form_params.pop('data', None)
                graphs_form_params['initial'] = {
                    'xselect': xind,
                    'yselect': yind
                }
                graphs_select_form = GraphSelectionForm(**graphs_form_params)
            # Recall y-axes stored in session (if any).
            y_axes = set(request.session.get(session_defs_key, [yind]))
            if submit_val.endswith('_add') and yind not in y_axes:
                y_axes.add(yind)
            if graphs_tab == 'tab_graphs_pie' or (
                    submit_val.endswith('_show') and len(y_axes) == 0):
                y_axes = set([yind])

            # Create form allowing the user to remove selected y-functions.
            remove_plots_list = []
            for yid in y_axes:
                remove_plots_list.append(
                    (graphs.graphs_y_axes[yid].description, yid))
            graph_remove_plots_form = GraphRemovePlotForm(
                remove_plots_list, request.POST)

            # Handle removing a y-function.
            if submit_val.endswith('_remove'):
                for yid in graph_remove_plots_form.get_to_del():
                    y_axes.remove(yid)
                    graph_remove_plots_form.fields.pop(str(yid))
            request.session[session_defs_key] = y_axes
            request.session[session_defs_key + '_x_ind'] = xind
            request.session[session_defs_key + '_y_ind'] = yind

            # Fetch graph data and pass it to the View template.
            graph_xfun_index = xind
            graph_data = get_graph_data(
                results, xfuns[xind], [graphs_y_axes[yind] for yind in y_axes])
        elif submit_val == 'tab_timeline':
            tab = 'timeline'

            timeline_form = TimelineVariableForm(request.POST)
            try:
                timeline_form_in_session = request.session[
                    'voyage_timeline_form_option']
            except Exception:
                timeline_form_in_session = None

            if not timeline_form.is_valid():
                timeline_form = timeline_form_in_session

            if timeline_form.is_valid():
                # If any choice passed, get chosen index and get selected tuple
                timeline_selected_var_index = int(
                    timeline_form.cleaned_data['variable_select'])
            else:
                # Otherwise, it's the first time when timeline is loaded
                timeline_form = TimelineVariableForm(
                    initial={'variable_select': '15'})
                timeline_selected_var_index = 15
            timeline_selected_tuple = voyage_timeline_variables[
                timeline_selected_var_index]

            request.session['voyage_timeline_form_option'] = timeline_form

            # Get set based on choice
            timeline_data = timeline_selected_tuple[2](
                results, timeline_selected_tuple[3])

            if request.POST and "download" in request.POST:
                return download_xls([[("Year", 1),
                                      (timeline_selected_tuple[1], 1)]],
                                    timeline_data)

            timeline_chart_settings['name'] = timeline_selected_tuple[1]
            if len(timeline_selected_tuple) > 4:
                # Include extra dict if exists
                items = list(timeline_chart_settings.items())
                items += list(timeline_selected_tuple[4].items())
                timeline_chart_settings = dict(items)
        elif submit_val in ('tab_maps', 'tab_animation'):
            tab = submit_val[4:]
            frame_from_year = int(
                request.POST.get('frame_from_year', voyage_span_first_year))
            frame_to_year = int(
                request.POST.get('frame_to_year', voyage_span_last_year))
            map_year = get_map_year(frame_from_year, frame_to_year)
        elif submit_val == 'map_ajax':
            map_ports = {}
            map_flows = {}

            def add_port(geo):
                result = geo and geo[0].show and geo[1].show and geo[2].show
                if result and not geo[0].pk in map_ports:
                    map_ports[geo[0].pk] = geo
                return result

            def add_flow(src, dst, embarked, disembarked):
                if embarked is None or disembarked is None:
                    return False
                if not (add_port(src) and add_port(dst)):
                    return False
                flow_key = int(src[0].pk) * 2147483647 + int(dst[0].pk)
                current = map_flows.get(flow_key)
                if current:
                    embarked += current[2]
                    disembarked += current[3]
                map_flows[flow_key] = (src[0].name, dst[0].name,
                                       embarked, disembarked)
                return True

            # Ensure cache is loaded.
            VoyageCache.load()
            all_voyages = VoyageCache.voyages
            missed_embarked = 0
            missed_disembarked = 0
            # Set up an unspecified source that will be used if the appropriate
            # setting is enabled
            geo_unspecified = CachedGeo(-1, -1, _('Africa, port unspecified'),
                                        0.05, 9.34, True, None)
            source_unspecified = (geo_unspecified, geo_unspecified,
                                  geo_unspecified)
            keys = get_pks_from_haystack_results(results)
            for pk in keys:
                voyage = all_voyages.get(pk)
                if voyage is None:
                    continue
                source = CachedGeo.get_hierarchy(voyage.emb_pk)
                if source is None and settings.MAP_MISSING_SOURCE_ENABLED:
                    source = source_unspecified
                destination = CachedGeo.get_hierarchy(voyage.dis_pk)
                add_flow(source, destination, voyage.embarked,
                         voyage.disembarked)
            if missed_embarked > 0 or missed_disembarked > 0:
                logging.getLogger('voyages').info(
                    'Missing flow: '
                    '(%d, %d)', missed_embarked, missed_disembarked)
            return render(request,
                          "voyage/search_maps.datatemplate", {
                              'map_ports': map_ports,
                              'map_flows': map_flows
                          },
                          content_type='text/javascript')
        elif submit_val == 'animation_ajax':
            VoyageCache.load()
            all_voyages = VoyageCache.voyages
            all_routes = VoyageRoutesCache.load()

            def animation_response():
                keys = get_pks_from_haystack_results(results)
                for pk in keys:
                    voyage = all_voyages.get(pk)
                    if voyage is None:
                        continue

                    def can_show(ph):
                        return ph is not None and ph[0].show

                    if not ok_to_show_animation(voyage, can_show):
                        continue

                    flag = VoyageCache.nations.get(
                        voyage.ship_nat_pk) or ''
                    ship_name = voyage.ship_name or ""
                    src = CachedGeo.get_hierarchy(voyage.emb_pk)
                    dst = CachedGeo.get_hierarchy(voyage.dis_pk)
                    route = all_routes.get(pk, ([],))[0]
                    yield ('{ '
                           '"voyage_id": ' + str(voyage.voyage_id) + ', '
                           '"source_name": "' + _(src[0].name) + '", '
                           '"source_lat": ' + str(src[0].lat) + ', '
                           '"source_lng": ' + str(src[0].lng) + ', '
                           '"destination_name": "' + _(dst[0].name) + '", '
                           '"destination_lat": ' + str(dst[0].lat) + ', '
                           '"destination_lng": ' + str(dst[0].lng) + ', '
                           '"embarked": ' + str(voyage.embarked) + ', '
                           '"disembarked": ' + str(voyage.disembarked) + ', '
                           '"year": ' + str(voyage.year) + ', '
                           '"ship_ton": ' + str(voyage.ship_ton or 0) + ', '
                           '"nat_id": ' + str(voyage.ship_nat_pk or 0) + ', '
                           '"ship_nationality_name": "' + _(flag) + '", '
                           '"ship_name": "' + str(ship_name) + '", '
                           '"route": ' + json.dumps(route) + ' '
                           '}')

            return HttpResponse('[' + ',\n'.join(animation_response()) + ']',
                                'application/json')
        elif submit_val == 'download_xls_current_page':
            page_num = request.POST.get('page_num') or 1
            return download_xls_page(results, int(page_num), results_per_page,
                                     display_columns, var_list)
        elif submit_val == 'delete_prev_query':
            prev_query_num = int(request.POST.get('prev_query_num'))
            qprev = request.session['previous_queries']
            qprev.remove(qprev[prev_query_num])
            request.session['previous_queries'] = qprev
            prev_queries_open = True

    if len(results) == 0:
        no_result = True

    # Paginate results to pages
    if tab == 'result':
        paginator = Paginator(results, results_per_page)
        pagins = paginator.page(int(current_page))
        request.session['current_page'] = current_page
        # Prepare paginator ranges
        (paginator_range,
         pages_range) = prepare_paginator_variables(paginator, current_page,
                                                    results_per_page)
        result_display = prettify_results(
            [x.get_stored_fields() for x in pagins], display_methods)
    else:
        pagins = None
        (paginator_range, pages_range) = (None, None)
        result_display = None
    # Customize result page
    if not request.session.exists(request.session.session_key):
        request.session.create()

    # Set up the initial column of display
    if 'result_columns' not in request.session:
        request.session['result_columns'] = get_new_visible_attrs(
            default_result_columns)
    if len(request.session.get('previous_queries', [])) > 10:
        request.session['previous_queries'] = request.session[
            'previous_queries'][:5]

    previous_queries = enumerate(
        map(prettify_var_list, request.session.get('previous_queries', [])))

    return render(
        request, "voyage/search.html", {
            'voyage_span_first_year': voyage_span_first_year,
            'voyage_span_last_year': voyage_span_last_year,
            'basic_variables': basic_variables,
            'general_variables': general_variables,
            'all_var_list': var_dict,
            'results': pagins,
            'result_data': result_data,
            'order_by_field': order_by_field,
            'sort_direction': sort_direction,
            'paginator_range': paginator_range,
            'pages_range': pages_range,
            'curpage': pagins,
            'no_result': no_result,
            'url_to_copy': search_url,
            'tab': tab,
            'options_results_per_page_form': results_per_page_form,
            'form_list': form_list,
            'time_frame_search_form': time_frame_form,
            'result_display': result_display,
            'basic_list_contracted': basic_list_contracted,
            'previous_queries': previous_queries,
            'prev_queries_open': prev_queries_open,
            'col_labels_list': collabels,
            'row_list': row_list,
            'table_stats_form': table_stats_form,
            'col_totals': col_totals,
            'extra_cols': list(range(extra_cols)),
            'num_col_labels_before': num_col_labels_before,
            'num_col_labels_total': num_col_labels_total,
            'num_row_labels': num_row_labels,
            'is_double_fun': is_double_fun,
            'inline_graph_png': inline_graph_png,
            'graph_remove_plots_form': graph_remove_plots_form,
            'graph_xfun_index': graph_xfun_index,
            'graphs_tab': graphs_tab,
            'graphs_select_form': graphs_select_form,
            'graph_data': graph_data,
            'timeline_data': timeline_data,
            'timeline_form': timeline_form,
            'timeline_chart_settings': timeline_chart_settings,
            'map_year': map_year
        })


def prettify_results(results, lookup_table):
    """
    Returns a list of dictionaries keyed by variable name, prettifies the value
    so that it displays properly Uses the lookup_table which has the methods to
    prettify the variable based on the value and the voyage id The lookup_table
    is gotten from the globals file, either from the display_methods or the
    display_methods_xls
    """
    # Results must be a list of dictionaries of variable name and value
    for i in results:
        idict = {}
        voyageid = int(i['var_voyage_id'])
        for varname, varvalue in list(i.items()):
            if varvalue:
                prettify_varvalue = lookup_table.get(
                    varname, default_prettifier(varname))
                if varvalue == u'[]':
                    varvalue = u''
                idict[varname] = prettify_varvalue(varvalue, voyageid)
            else:
                idict[varname] = None
        yield idict


def _get_all(model):
    try:
        return list(model.objects.all())
    except Exception:
        return []


class _ChoicesCache:
    nations = _get_all(Nationality)
    particular_outcomes = _get_all(ParticularOutcome)
    slaves_outcomes = _get_all(SlavesOutcome)
    owner_outcomes = _get_all(OwnerOutcome)
    resistances = _get_all(Resistance)
    captured_outcomes = _get_all(VesselCapturedOutcome)
    rigs = _get_all(RigOfVessel)


def get_choices(varname):
    """
    Retrieve a list of two-tuple items for select boxes depending on the model
    :param varname variable name:
    :return: nested list of choices/options for that variable
    """
    choices = []
    if varname in ['var_nationality']:
        for nation in _ChoicesCache.nations:
            if set({'/', 'Other (specify in note)'}).isdisjoint(nation.label):
                choices.append((nation.value, _(nation.label)))
    elif varname in ['var_imputed_nationality']:
        for nation in _ChoicesCache.nations:
            # imputed flags
            if nation.label in list_imputed_nationality_values:
                choices.append((nation.value, _(nation.label)))
    elif varname in ['var_outcome_voyage']:
        for outcome in _ChoicesCache.particular_outcomes:
            choices.append((outcome.value, _(outcome.label)))
    elif varname in ['var_outcome_slaves']:
        for outcome in _ChoicesCache.slaves_outcomes:
            choices.append((outcome.value, _(outcome.label)))
    elif varname in ['var_outcome_owner']:
        for outcome in _ChoicesCache.owner_outcomes:
            choices.append((outcome.value, _(outcome.label)))
    elif varname in ['var_resistance']:
        for outcome in _ChoicesCache.resistances:
            choices.append((outcome.value, _(outcome.label)))
    elif varname in ['var_outcome_ship_captured']:
        for outcome in _ChoicesCache.captured_outcomes:
            choices.append((outcome.value, _(outcome.label)))
    elif varname == 'var_rig_of_vessel':
        for rig in _ChoicesCache.rigs:
            choices.append((rig.value, _(rig.label)))
    return choices


def put_other_last(lst):
    others = [
        x for x in lst
        if 'other' in x['text'].lower() or ''
        'unspecified' in x['text'].lower() or '???' in x['text']
    ]
    for rem in others:
        lst.remove(rem)
        lst.append(rem)
    return lst


# , place_visible=[], region_visible=[], area_visible=[], place_selected=[],
# region_selected=[], area_selected=[]):
def get_nested_list_places(varname, nested_places, selected_places=None):
    """
    Retrieve a nested list of places sorted by broad region (area) and then
    region
    :param varname:
    returns a tuple of the nested choices and the choices
    """
    nested_choices = []
    flat_choices = []

    select = [int(i) for i in selected_places] if selected_places else []

    for area, regs in list(nested_places.items()):
        area_content = []
        is_area_selected = True
        for reg, places in list(regs.items()):
            reg_content = []
            is_selected = True
            for place in places:
                if place.value not in select:
                    is_selected = False
                    is_area_selected = False
                    # Why did I have this here?


#                if place.place == "???":
#                    continue
# Selected does not need to be part of this, since the form dict will have a
# list of places selected that the template checks
                reg_content.append({
                    'id': 'id_' + varname + '_2_' + str(place.pk),
                    'text': place.place,
                    'value': place.value
                })
                flat_choices.append((place.value, place.place))
            reg_content = put_other_last(reg_content)
            area_content.append({
                'id': 'id_' + varname + '_1_' + str(reg.pk),
                'text': reg.region,
                'choices': reg_content,
                'is_selected': is_selected,
                'value': reg.value
            })
        area_content = put_other_last(area_content)
        nested_choices.append({
            'id': 'id_' + varname + '_0_' + str(area.pk),
            'text': area.broad_region,
            'choices': area_content,
            'is_selected': is_area_selected,
            'value': area.value
        })
    nested_choices = put_other_last(nested_choices)

    # Check if visible parameters have been passed, if so filter
    return nested_choices, flat_choices


def prepare_paginator_variables(paginator, current_page, results_per_page):
    """
    Function prepares set of paginator links for template.

    :param paginator: Paginator which links are calculating for
    :param current_page: Current page serves on search site
    """

    paginator_range = []
    pages_range = []
    last_saved_index = 0

    # Prepare page numbers
    for i in paginator_range_factors:

        # Get last inserted index
        try:
            last = paginator_range[-1]
        except IndexError:
            last = 0

        # If page number would be greater than max page number,
        # return, since this is the end of page paginator ranges
        if last_saved_index >= len(paginator.object_list):
            continue
        # Index can't be less than '1'
        if int(current_page) + i < 1:
            paginator_range.append(last + 1)
            last_saved_index = ((last + 1) * results_per_page)
        else:
            # Index has to be always +1 from the last one
            # (e.g. current_page is '1')
            if int(current_page) + i <= last:
                paginator_range.append(last + 1)
                last_saved_index = ((last + 1) * results_per_page)
            # Otherwise, just increment current_page
            else:
                paginator_range.append(int(current_page) + i)
                last_saved_index = ((int(current_page) + i) * results_per_page)

    # Prepare results range
    pages_range.append(
        int(current_page) * results_per_page - results_per_page + 1)
    if (int(current_page) * results_per_page) > len(paginator.object_list):
        pages_range.append(len(paginator.object_list))
    else:
        pages_range.append(int(current_page) * results_per_page)

    return paginator_range, pages_range


def search_var_dict(var_name):
    for i in var_dict:
        if i['var_name'] == var_name:
            return i
    return None


# Automatically fetch fields that should be sorted differently, either
# by using a translated version or a plain text version of tokenized fields.

index = VoyageIndex()
plain_text_suffix = '_plaintext'
plain_text_suffix_list = [
    f[:-len(plain_text_suffix)]
    for f in list(index.fields.keys())
    if f.endswith(plain_text_suffix)
]
translate_suffix = '_lang_en'
translated_field_list = [
    f[:-len(translate_suffix)]
    for f in list(index.fields.keys())
    if f.endswith(translate_suffix)
]


def perform_search(query_dict,
                   date_filters,
                   order_by_field='var_voyage_id',
                   sort_direction='asc',
                   lang='en'):
    """
    Perform the actual query towards SOLR
    :param query_dict:
    :param date_filters:
    :return:
    """
    if order_by_field in plain_text_suffix_list:
        order_by_field += '_plaintext_exact'
    elif order_by_field in translated_field_list:
        order_by_field += '_lang_' + lang + '_exact'
    if sort_direction == 'desc':
        order_by_field = '-' + order_by_field
    results = get_voyages_search_query_set().filter(
        **query_dict).order_by(order_by_field)
    # Date filters
    return date_filter_query(date_filters, results)


# TODO: remove this function


def date_filter_query(date_filters, results):
    """
    Further filter the results passed in by excluding those that have months in
    date_filtered list (deselected)
    :param date_filters:
    :param results:
    :return:
    """
    if date_filters:
        for var_filter in date_filters:
            l_months = []
            tmp_query = dict()
            for month in var_filter['deselected_months']:
                l_months.append("-" + month + "-")

            tmp_query[var_filter['varname'] + "__in"] = l_months
            results = results.exclude(**tmp_query)

    return results


def format_date(year, month):
    """
    Format the passed year month to a YYYY,MM string
    :param year:
    :param month:
    :return:
    """
    if month == "":
        month = 1
    return date(int(year), int(month), 1)
    # return "%s,%s" % (str(year).zfill(4), str(month).zfill(2))


def get_new_visible_attrs(list_column_varnames):
    """
    :param list_column_varnames: a list of variable names (short_name)
    :return: a list of tuples containing short names and full names of
    variables from the passed parameter
    """
    result_columns = []
    for column in list_column_varnames:
        for item in var_dict:
            if item['var_name'] == column:
                result_columns.append(
                    [item['var_name'], item['var_full_name']])
    return result_columns


def download_xls_page(results, current_page, results_per_page, columns,
                      var_list):
    # Download only current page
    if current_page != -1:
        paginator = Paginator(results, results_per_page)
        curpage = paginator.page(current_page)
        res = [x.get_stored_fields() for x in curpage.object_list]
    else:
        res = list(results.values(*[x[0] for x in columns]).all())
    pres = prettify_results(res, display_methods_xls)

    response = HttpResponse(content_type='application/vnd.ms-excel')
    response['Content-Disposition'] = 'attachment; filename="data.xls"'
    wb = xlwt.Workbook(encoding='utf-8')
    ws = wb.add_sheet("data")
    # TODO: add query to download
    if len(var_list.get('used_variable_names', [])) > 0:
        vartxt = "; ".join(
            [i[0] + " " + i[1] for i in prettify_var_list(var_list)])
        ws.write(0, 0, label=vartxt)
    elif current_page == -1:
        ws.write(0, 0, label='All Records')
    else:
        ws.write(0, 0, label='Current Page')
    for idx, column in enumerate(columns):
        ws.write(1, idx, label=get_spss_name(column[0]))

    for idx, item in enumerate(pres):
        for idy, column in enumerate(columns):
            data = item[column[0]]
            if data is None:
                ws.write(idx + 2, idy, label="")
            elif isinstance(data, (int, float)):
                ws.write(idx + 2, idy, label=data)
            else:
                ws.write(idx + 2, idy, label=data.encode("utf-8"))

    wb.save(response)

    return response


def csv_stats_download(request):
    """
    Renders a downloadable csv file of summary statistics
    :param request:
    :param page:
    :return:
    """
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="data.csv"'

    writer = csv.writer(response)
    query_dict = request.session.get('voyage_last_query')
    if query_dict:
        results = perform_search(
            query_dict,
            request.session.get('voyage_last_query_date_filters', []),
            lang=request.LANGUAGE_CODE)

        if len(results) == 0:
            results = []
    else:
        results = get_voyages_search_query_set().order_by('var_voyage_id')

    # Write headers
    row = []
    for columnname in summary_statistics_columns:
        row.append(columnname)
    writer.writerow(row)

    for row in retrieve_summary_stats(results):
        writer.writerow(row)
    return response


def get_spss_name(var_short_name):
    """
    Retrieves a variable spss name based on its django name
    :param var_short_name:
    :return:
    """
    for var in var_dict:
        if var['var_name'] == var_short_name:
            return var['spss_name']
    return None


def retrieve_summary_stats(results):
    """
    Return a list of summary statistics
    :param results:
    :return:
    """
    tmp_list = []
    for item in summary_statistics:
        tmp_row = [
            item['display_name'],
        ]
        stats = results.stats(item['var_name']).stats_results().get(
            item['var_name'])

        if item['has_total'] and stats:
            tmp_row.append(int(stats['sum']))
        else:
            tmp_row.append("")

        # Number of voyages
        count = 0
        if stats:
            count = int(stats['count'])
            tmp_row.append(count)
        else:
            tmp_row.append("")

        if stats and count > 0:
            if item['is_percentage']:
                # Average
                tmp_row.append(str(round(stats['mean'] * 100, 1)) + "%")
                # Standard deviation
                tmp_row.append(str(round(stats['stddev'] * 100, 1)) + "%")
            else:
                tmp_row.append(round(stats['mean'], 1))
                tmp_row.append(round(stats['stddev'], 1))
        else:
            tmp_row.append("")
            tmp_row.append("")

        tmp_list.append(tmp_row)
    return tmp_list


def remove_empty_columns(idx, row_list, collabels, col_totals):
    finish = True

    # Check if any has to be removed
    for i in range(idx, len(col_totals) - 1):
        if col_totals[i] == 0:
            idx = i
            finish = False
            break

    # If not, finish
    if finish:
        return

    # Remove column in row_list (values)
    for i in row_list:
        del i[1][idx]

    # Remove total column
    del col_totals[idx]

    # Remove port and any parents if needed
    # Port
    current_index = 1
    for ind, value in enumerate(collabels[-1]):
        if value[1] != 1:
            continue
        if current_index == idx + 1:
            del collabels[-1][ind]
            break
        current_index += 1

    # If there is more (parents), check then and remove if necessary
    # Start from first parent and go to grandparents if necessary
    depth = len(collabels) - 2

    # As long as parent exists
    while depth > -1:
        current_list = collabels[depth]

        # val - keep adding values of parents
        # count_items - counting how many items (parents) already checked
        val = 1
        count_items = 0

        for i, value in enumerate(current_list):
            # Check only items that are displayed (>1)
            if value[1] <= 0:
                continue
            # Increment counter of items
            count_items += 1

            # If wanted idx is in between of the current val and less than
            # val+next, found
            if val <= current_index < val + value[1]:

                # Decrement current parent value
                value[1] -= 1

                # If 0, remove it
                if value[1] == 0:
                    del collabels[depth][i]
                depth -= 1

                # Go to the next parents if needed
                break
            val += value[1]

    remove_empty_columns(idx + 1, row_list, collabels, col_totals)


def get_permanent_link(request):
    """
    Obtain a permanent link for the current search query.
    :param request: The request containing the search quer
    :param request: The request containing the search query.
    :return: A permanent URL link for that exact query.
    """
    saved_query = SavedQuery()
    return saved_query.get_link(request, 'restore_v_permalink')


def restore_permalink(_, link_id):
    """Redirect the page with a URL param"""
    return redirect("/voyage/database#searchId=" + link_id)
